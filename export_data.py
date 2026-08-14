#!/usr/bin/env python3
"""Regenerate data.json and data.public.json from Lab_Papers_Dashboard.xlsx.

Usage: python3 export_data.py [path/to/Lab_Papers_Dashboard.xlsx]
"""
import json
import os
import sys
from datetime import datetime, date, timezone

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: run `pip install openpyxl` first.")

# Fields on a paper that the update-request Action (scripts/apply_update.py)
# can change. These are the ones reconciliation may preserve from the
# existing data.json instead of overwriting with the spreadsheet's values —
# see reconcile_papers() below.
ISSUE_EDITABLE_FIELDS = [
    "stage", "priority", "owner", "attempts", "currentVenue", "deadline",
    "daysLeft", "latestDecision", "status", "notes", "lastUpdated",
    "submissions",
]

PAPERS_SHEET = "\U0001F4C4 Papers Tracker"
SUBMISSIONS_SHEET = "\U0001F4E8 Submissions Log"
TEAM_SHEET = "\U0001F465 Team Directory"

# These two sheets don't exist in the workbook yet — see the "Meetings /
# resources schema" note in README.md for the exact columns to add when
# ready. Until then, load_meetings()/load_resources() below just return
# empty results for every paper; nothing here requires them to exist.
MEETINGS_SHEET = "\U0001F4C5 Meetings Log"
RESOURCES_SHEET = "\U0001F517 Resources"

PAPERS_HEADER_ROW = 3
SUBMISSIONS_HEADER_ROW = 3
TEAM_HEADER_ROW = 3
MEETINGS_HEADER_ROW = 3
RESOURCES_HEADER_ROW = 3

# Placeholder shared-inbox address used on data.public.json whenever a
# paper's owner hasn't been explicitly marked OK to show their real email
# (Team Directory's "Public Contact OK" column) — see build_public_dataset()
# below. Change this to your lab's real shared inbox before publishing.
GENERIC_LAB_EMAIL = "criticaldata-lab@mit.edu"


def cell_value(v):
    if isinstance(v, (datetime, date)):
        return v.date().isoformat() if isinstance(v, datetime) else v.isoformat()
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def sheet_rows(ws, header_row):
    headers = [c.value for c in ws[header_row]]
    for row in ws.iter_rows(min_row=header_row + 1):
        values = [cell_value(c.value) for c in row]
        if not any(values):
            continue
        yield dict(zip(headers, values))


def as_int(v, default=0):
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def parse_list_cell(v):
    """'clinical ML, causal inference, EHR data' -> ['clinical ML', 'causal inference', 'EHR data']"""
    if not v:
        return []
    return [part.strip() for part in str(v).split(",") if part.strip()]


def parse_bool_cell(v):
    if v is None:
        return False
    return str(v).strip().lower() in ("yes", "true", "y", "1")


def parse_meeting_dt(s):
    """Parse a meeting date/time cell into a datetime for chronological
    sorting. Accepts a full ISO datetime or a bare date (treated as
    midnight) — cell_value() already normalizes Excel datetimes to ISO
    strings, so this only needs to handle those two shapes."""
    if not s:
        return None
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        try:
            return datetime.combine(date.fromisoformat(s[:10]), datetime.min.time())
        except ValueError:
            return None


def load_meetings(wb, now=None):
    """Returns {paper title: {"next": {...}|None, "past": [...]}}.

    One row per meeting in the Meetings Log sheet (mirrors the Submissions
    Log pattern: one row per attempt, joined by paper title). Whichever
    meeting is soonest in the future becomes `next`; everything else in the
    past becomes `past`, most recent first. A paper with no meetings logged
    simply doesn't appear in the returned dict.
    """
    if MEETINGS_SHEET not in wb.sheetnames:
        return {}
    now = now or datetime.now()
    by_paper = {}
    for row in sheet_rows(wb[MEETINGS_SHEET], MEETINGS_HEADER_ROW):
        title = row.get("Paper")
        if not title:
            continue
        by_paper.setdefault(title, []).append({
            "date": row.get("Date/Time"),
            "link": row.get("Link"),
            "notes": row.get("Notes"),
        })

    result = {}
    for title, meetings in by_paper.items():
        future = [m for m in meetings if (parse_meeting_dt(m["date"]) or datetime.min) >= now]
        past = [m for m in meetings if m not in future]
        future.sort(key=lambda m: parse_meeting_dt(m["date"]) or datetime.min)
        past.sort(key=lambda m: parse_meeting_dt(m["date"]) or datetime.min, reverse=True)
        result[title] = {
            "next": ({"date": future[0]["date"], "link": future[0]["link"]} if future else None),
            "past": past,
        }
    return result


def load_resources(wb):
    """Returns {paper title: [{"label", "url"}, ...]}. One row per
    resource, same joined-by-title pattern as Meetings/Submissions Log."""
    if RESOURCES_SHEET not in wb.sheetnames:
        return {}
    by_paper = {}
    for row in sheet_rows(wb[RESOURCES_SHEET], RESOURCES_HEADER_ROW):
        title = row.get("Paper")
        url = row.get("URL")
        if not title or not url:
            continue
        by_paper.setdefault(title, []).append({"label": row.get("Label") or url, "url": url})
    return by_paper


def public_stage_label(paper):
    """A plain-language stage label safe to show prospective members —
    deliberately coarser than the internal emoji status badge, which can
    say things like "Needs Attention" or "Overdue" that read badly out of
    context and aren't this audience's business anyway."""
    if paper.get("publishedDate"):
        return "Published"
    if (paper.get("attempts") or 0) > 0:
        return "Submitted, under review"
    stage = paper.get("stage")
    if stage == "Drafting":
        return "In progress (drafting)"
    if stage == "Internal Review":
        return "In progress (internal review)"
    if stage == "Idea":
        return "Early stage / idea"
    if stage == "On Hold":
        return "On hold"
    return "Getting started"


# The complete set of fields data.public.json's paper objects may ever
# carry. This is enforced two ways: build_public_dataset() below only ever
# constructs objects with exactly these keys (an ALLOWLIST — everything on
# the internal `papers` dict is excluded by default, not included by
# default), and worker-phase2-paused aside, a Playwright test asserts this
# exact key set on every run — see the "field-allowlist" test in
# test/ for what fails if this ever drifts.
PUBLIC_PAPER_FIELDS = ["id", "title", "abstract", "tags", "skillsNeeded", "stage", "openToNewMembers", "contact"]


def build_public_dataset(papers, team):
    """Builds data.public.json's papers list for the public Project
    Discovery page (discover.html).

    Two safety properties, both deliberate:
      1. Only papers with openToNewMembers=True are included AT ALL — this
         is filtered here, server-side, at export time, not left to
         discover.html's JS to filter client-side. A paper the lab hasn't
         marked open never appears in the file discover.html fetches, so
         there's nothing for a curious dev-tools user to find.
      2. Every field below is an explicit, positive inclusion. Nothing
         else on a paper — venue, deadline, notes, meeting links, draft
         links, resources, submission history, priority, or a raw personal
         email — is ever written here, because build_public_dataset()
         simply never reads those fields, not because they're filtered out
         after the fact.

    Contact info: an owner's real email is only used if their Team
    Directory row has "Public Contact OK" set — otherwise (the default)
    every paper's contact is the generic lab inbox, never a personal
    address pulled raw off the roster.
    """
    team_by_name = {m["name"]: m for m in team}
    public_papers = []
    for p in papers:
        if not p.get("openToNewMembers"):
            continue
        owner = p.get("owner")
        member = team_by_name.get(owner) if owner else None
        first_name = owner.split(" ")[0] if owner else None
        if member and member.get("publicContactOk") and member.get("email"):
            email = member["email"]
        else:
            email = GENERIC_LAB_EMAIL
        public_papers.append({
            "id": p["id"],
            "title": p["title"],
            "abstract": p.get("abstract"),
            "tags": p.get("tags") or [],
            "skillsNeeded": p.get("skillsNeeded") or [],
            "stage": public_stage_label(p),
            "openToNewMembers": True,
            "contact": {"name": first_name, "email": email},
        })
    return public_papers


def reconcile_papers(fresh_papers, previous_papers_by_id, spreadsheet_modified):
    """Merge freshly-exported papers with any changes already applied via
    the update-request Action, so re-running this script doesn't silently
    clobber someone's issue-driven update just because the spreadsheet
    hasn't caught up yet.

    The signal used is wall-clock time: each paper's `updatedViaIssue.at`
    timestamp (stamped by apply_update.py) is compared against the
    spreadsheet file's own last-saved time (from its docProps metadata).
    Whichever is more recent wins:

      - issue edit newer than the spreadsheet -> keep the JSON's values for
        the fields the Action can touch (stage/priority/owner/status/
        submissions/etc — see ISSUE_EDITABLE_FIELDS), and print a warning
        so whoever runs this knows to catch the spreadsheet up too.
      - spreadsheet saved more recently -> assume a human has since caught
        the spreadsheet up (or overridden it on purpose); use the fresh
        spreadsheet values as normal, and drop the now-stale
        `updatedViaIssue` marker.

    Returns the reconciled paper list; prints warnings for any preserved
    papers as a side effect.
    """
    reconciled = []
    for paper in fresh_papers:
        previous = previous_papers_by_id.get(paper["id"])
        marker = (previous or {}).get("updatedViaIssue")

        if previous and marker and marker.get("at"):
            issue_at = datetime.fromisoformat(marker["at"])
            if issue_at.tzinfo is None:
                issue_at = issue_at.replace(tzinfo=timezone.utc)
            if issue_at > spreadsheet_modified:
                print(
                    f"  keeping issue-applied values for {paper['id']} "
                    f"({paper['title'][:60]!r}) — update-request #{marker.get('issueNumber')} "
                    f"was applied {marker['at']}, after the spreadsheet was last saved. "
                    "Update the spreadsheet to match if you want it to stick."
                )
                merged = dict(paper)
                for field in ISSUE_EDITABLE_FIELDS:
                    merged[field] = previous.get(field)
                merged["updatedViaIssue"] = marker
                reconciled.append(merged)
                continue

        reconciled.append(paper)
    return reconciled


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "Lab_Papers_Dashboard.xlsx"
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    submissions_by_paper = {}
    for row in sheet_rows(wb[SUBMISSIONS_SHEET], SUBMISSIONS_HEADER_ROW):
        title = row.get("Paper")
        if not title:
            continue
        submissions_by_paper.setdefault(title, []).append({
            "attempt": as_int(row.get("Attempt #"), None),
            "venue": row.get("Venue / Journal or Conference"),
            "submittedDate": row.get("Submitted Date"),
            "responseDeadline": row.get("Response Deadline"),
            "decision": row.get("Decision"),
            "decisionDate": row.get("Decision Date"),
            "notes": row.get("Notes"),
        })
    for subs in submissions_by_paper.values():
        subs.sort(key=lambda s: (s["attempt"] is None, s["attempt"]))

    meetings_by_paper = load_meetings(wb)
    resources_by_paper = load_resources(wb)

    papers = []
    for row in sheet_rows(wb[PAPERS_SHEET], PAPERS_HEADER_ROW):
        title = row.get("Paper Title")
        if not title:
            continue
        meetings = meetings_by_paper.get(title, {"next": None, "past": []})
        papers.append({
            "id": row.get("ID"),
            "title": title,
            "stage": row.get("Stage"),
            "status": row.get("Status"),
            "priority": row.get("Priority"),
            "targetVenue": row.get("Target Venue\n(pre-submission)"),
            "targetDeadline": row.get("Target Deadline\n(pre-submission)"),
            "attempts": as_int(row.get("Attempts")),
            "currentVenue": row.get("Current Venue"),
            "deadline": row.get("Deadline"),
            "daysLeft": as_int(row.get("Days Left"), None),
            "latestDecision": row.get("Latest Decision"),
            "owner": row.get("Owner"),
            "authors": row.get("Authors"),
            "finalFile": row.get("Final File"),
            "publishedDate": row.get("Published Date"),
            "progressPct": as_int(row.get("Progress %"), None),
            "lastUpdated": row.get("Last Updated"),
            "githubRepo": row.get("GitHub Repo"),
            "notes": row.get("Notes"),
            "submissions": submissions_by_paper.get(title, []),
            "nextMeeting": meetings["next"],
            "pastMeetings": meetings["past"],
            "currentDraftLink": row.get("Current Draft Link"),
            "resources": resources_by_paper.get(title, []),
            "abstract": row.get("Abstract"),
            "tags": parse_list_cell(row.get("Tags")),
            "skillsNeeded": parse_list_cell(row.get("Skills Needed")),
            "openToNewMembers": parse_bool_cell(row.get("Open to New Members")),
        })

    team = []
    for row in sheet_rows(wb[TEAM_SHEET], TEAM_HEADER_ROW):
        name = row.get("Name")
        # Skip instructional footer rows (e.g. "Add every additional lab
        # member..."), which have text in the Name column but no Role.
        if not name or not row.get("Role"):
            continue
        team.append({
            "name": name,
            "role": row.get("Role"),
            "focusArea": row.get("Focus Area"),
            "currentPapers": as_int(row.get("Current Papers")),
            "email": row.get("Email"),
            "availability": row.get("Availability"),
            "publicContactOk": parse_bool_cell(row.get("Public Contact OK")),
        })

    previous_papers_by_id = {}
    if os.path.exists("data.json"):
        with open("data.json", encoding="utf-8") as f:
            try:
                previous_papers_by_id = {p["id"]: p for p in json.load(f).get("papers", [])}
            except (json.JSONDecodeError, AttributeError):
                previous_papers_by_id = {}

    spreadsheet_modified = wb.properties.modified or wb.properties.created
    if spreadsheet_modified.tzinfo is None:
        spreadsheet_modified = spreadsheet_modified.replace(tzinfo=timezone.utc)

    papers = reconcile_papers(papers, previous_papers_by_id, spreadsheet_modified)

    data = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "papers": papers,
        "team": team,
    }

    with open("data.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"Wrote data.json: {len(papers)} papers, {len(team)} team members.")

    public_papers = build_public_dataset(papers, team)
    public_data = {
        "generatedAt": data["generatedAt"],
        "papers": public_papers,
    }
    with open("data.public.json", "w", encoding="utf-8") as f:
        json.dump(public_data, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"Wrote data.public.json: {len(public_papers)} paper(s) open to new members.")


if __name__ == "__main__":
    main()
